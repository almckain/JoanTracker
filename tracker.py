from ast import operator
from webbrowser import get
from openpyxl.utils import get_column_letter
from datetime import date, datetime, timedelta
from openpyxl import Workbook, load_workbook
import operator
import matplotlib.pyplot as plt

#Location of spreadsheet
wb = load_workbook("/Users/aaronmckain/Desktop/JoanTrackerProj/unitTracker.xlsx")
ws = wb.active

#Empty list
room_locations = []

room_and_averages = {
    #Empty dictionary
}

def calculate_sum(num_of_times):
    sum_of_days = timedelta(days=0)
    
    for i in range(0, len(num_of_times) - 1):
        difference = num_of_times[i] - (num_of_times[i + 1])
        sum_of_days = sum_of_days + difference
        
        """ Debugging print statement and variables
        x = num_of_times[i]
        y = num_of_times[i + 1]
        print(f"\t\t{difference} = {x} - {y}")
        """
    average = sum_of_days / len(num_of_times)
    return average

def sort_room_averages(averages_dictionary):
    print("I dont know how to sort a dictionary lmao")
        



#Stores the cell letters in a list
for row in range(1,2): #rows 1-2
    for col in range(1, 46): #col 1-40
        room_locations.append(get_column_letter(col))
        for i in room_locations:
            print(i, end = " ")
        print("\n") 



#Cycles through each cell
for cell in room_locations:
    #temporary list that store the times from each room; clears after use
    list_of_times = []
    for i in range(2,39):
        time = ws[cell + str(i)]
        #stores cell data in the temporary list if return is not null
        if(time.value != None):
            list_of_times.append((time.value).date())
        
    #Makes the temporary list sorted from earliest to latest      
    list_of_times.reverse()

    #Calls calculate average
    average_change = calculate_sum(list_of_times)
    print("The average days between changes is " + str(average_change.days))
    
    #Assigns room to key
    room = ws[cell + str(1)]
    
    #Assigns average to value
    room_and_averages[str(room.value)] = int(average_change.days)

    #clears the list for the next room
    list_of_times.clear()
    print("\n")

#Sorts dictionary
sorted_dict = dict(sorted(room_and_averages.items(), key=operator.itemgetter(1)))
#sorted_dict = sorted(room_and_averages.items())
#View dictionary
#print(sorted_dict)



rooms = list(sorted_dict.keys())
changes = list(sorted_dict.values())

print("\n")
print(type(rooms[3]))
print(type(changes[1]))
plt.title("Days Between Joan Changes")
plt.ylabel("Room Number")
plt.xlabel("Days")
plt.barh(rooms, changes)
plt.tight_layout()

plt.show()








